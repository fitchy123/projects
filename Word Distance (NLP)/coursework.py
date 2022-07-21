##python version 3.8.5
##problem 1 and 2 - takes about 2 min 40 secs seconds to run on quite slow wifi
##I have combined these problems for efficiency
import openpyxl
import csv
import requests
from requests.exceptions import HTTPError
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

start_time = time.time()

def modern_scraper(just_article):
    article_string = ""
    #get the article header
    article_heading = just_article.find("h1")
    article_string += article_heading.text + "\n"
    #get all relevant text, this includes figure captions and the main article contents
    #preserve the order
    if just_article.findAll(class_=["ssrcss-1rnnz6t-StyledFigureCaption e34k3c21", 
    "ssrcss-uf6wea-RichTextComponentWrapper e1xue1i83", "ssrcss-qozapo-StyledHeading e1fj1fc10"]):
        text_list = just_article.findAll(class_=["ssrcss-1rnnz6t-StyledFigureCaption e34k3c21", 
    "ssrcss-uf6wea-RichTextComponentWrapper e1xue1i83", "ssrcss-qozapo-StyledHeading e1fj1fc10"])
    else:
        text_list = just_article.findAll(class_="ssrcss-3z08n3-RichTextContainer e5tfeyi2")

    prev = article_string
    for txt in text_list:
        act_txt = txt.text
        #ensure there are no duplicate bits of text
        if article_string[-1] == "\n":
            prev = article_string[-len(act_txt)-1:]
        else:
            prev = article_string[-len(act_txt):]
        #ensure irrelevant text like links to other parts of the website or caption labels are discarded
        if "image caption" in act_txt or "media caption" in act_txt or "/newsround/" in act_txt:
            article_string += act_txt[13:] + ".\n"
        elif "Facebook, Twitter and Instagram" in act_txt or "Facebook, Instagram and Twitter" in act_txt:
            continue
        elif prev.rstrip() == act_txt.rstrip():
            continue
        else:
            article_string += act_txt + "\n"
    
    return article_string

def middle_scraper(just_article):
    article_string = " "
    #get the article header
    if just_article.find(class_="mxb"):
        article_heading = just_article.find(class_="mxb")
        article_string += article_heading.text.strip() + "\n"
    #isolate the story body, preserve the order
    story_body = just_article.find(class_="storybody")
    text_list = story_body.select("p, div[class^='mva']")
    prev = article_string
        
    for txt in text_list:
        #discard links to other parts of website
        if txt.find("a"):
            continue
        act_text = txt.text
        #ensure there are no duplicate bits of text
        if article_string[-1] == "\n":
            prev = article_string[-len(act_text)-1:]
        else:
            prev = article_string[-len(act_text):]
        if prev.rstrip() == act_text.rstrip():
            continue
        article_string += act_text 
    
    return article_string

def old_scraper(just_article):
    article_string = ""
    #get article header
    article_heading = just_article.find(class_="mxb")
    article_string += article_heading.text.strip() + "\n"
    #isolate article body, preserve order
    article_body = just_article.find("td", valign="top", width="416")
    text_list = article_body.findAll(["b", "p"])
    prev = article_string
    for txt in text_list:
        act_text = txt.text
        #ensure no duplicate bits of text
        if article_string[-1] == "\n":
            prev = article_string[-len(act_text)-1:]
        else:
            prev = article_string[-len(act_text):]
        #discard links to other parts of the website
        if "LINKS TO MORE" in act_text:
            continue
        elif prev.rstrip() == act_text.rstrip():
            continue
        article_string += act_text

    return article_string

def older_scraper(just_article):
    article_string = ""
    #get article header
    article_heading = just_article.find(class_="headlinestory")
    article_string += article_heading.text
    #isolate article body
    article_body = just_article.find(class_="bodytext")
    article_string += article_body.text
    return article_string

def oldest_scraper(just_article):
    article_string = " "
    #get article 
    just_article = just_article.find("td", valign="top", width="328")
    #get article text, preserve order
    article_txt = just_article.findAll(["b", "p"])
    for txt in article_txt:
        act_text = txt.text
        #ensure no duplicate bits of text
        if article_string[-1] == "\n":
            prev = article_string[-len(act_text)-1:]
        else:
            prev = article_string[-len(act_text):]
        if prev.rstrip() == act_text.rstrip():
            continue
        #discard irrelevant bits of text
        elif "Back to top" in act_text or "Advanced options" in act_text or "Sci/Tech" in act_text:
            continue
        article_string += act_text
    return article_string

def scraper(article_link):
    article_retries = 0
    for _ in range(2):   
        try:
            #request resources from article url
            response = requests.get(article_link, timeout=2.0)
            #raise a HTTPError if HTTP request has unsucessful status code
            response.raise_for_status()
        #deal with HTTP errors
        except HTTPError as httperr:
            print(f"HTTP error: {httperr}")
            print("Retrying up to 2 times, currently on retry: ", article_retries+1)
            article_retries += 1
            continue
        #dealing with non HTTP errors
        except Exception as error:
            print(f"Other error: {error}")
            print("Retrying up to 2 times, currently on retry: ", article_retries+1)
            article_retries += 1
            continue
        break
    
    #if an error occurs, increment article retries and try again, if 2 errors occur in a row move on to next article
    if article_retries > 1:
        return 

    #parse the unicode content of article with lxml parser
    parsed_article = BeautifulSoup((response.text), "lxml")

    #extract relevant info, articles from different time periods have different formats
    #I have made a scraping function for each of these formats
    if parsed_article.find("article"):
        only_article = parsed_article.find("article")
        article_str = modern_scraper(only_article)
    elif parsed_article.find(class_="storycontent"):
        only_article = parsed_article.find(class_="storycontent")
        article_str = middle_scraper(only_article) 
    elif parsed_article.find("td", valign="top", width="416"):
        article_str = old_scraper(parsed_article)
    elif parsed_article.find(class_="headlinestory"):
        article_str = older_scraper(parsed_article)
    elif parsed_article.find("font", face="Arial, Helvetica"):
        article_str = oldest_scraper(parsed_article)
    else:
        #notify if article is not a format seen before
        print("Invalid format, moving on to next article")
        return

    return article_str 


#open excel file/workbook and create workbook object
filename = r"keywords.xlsx"
excel_file = openpyxl.load_workbook(filename)
excel_keywords = excel_file.active

#list of keywords
keywords = []

#loop through 1st column (except the title of the column) to read in keywords
for row_ in range(2, excel_keywords.max_row + 1):
    word = excel_keywords.cell(row = row_, column = 1)
    keywords.append(word.value)

#close excel file
excel_file.close()

#Initialise list of search urls for each keyword(s)
#The links to actual 100 articles relevant to these keywords will be scraped from these urls
keyword_urls = []
base_url = "https://www.bbc.co.uk/search?q="

#create list of search urls for each keyword
for keyword in keywords:
    keyword_url = base_url
    count = 0
    for word in keyword.split():
        #if there is more than one word in the keyword(s) add a "+" so the url is valid
        if count > 0:
            keyword_url += "+"
        keyword_url += word 
        count += 1
    keyword_urls.append(keyword_url + "&page=")

#open csv file that article data will be saved to
#this data structure will have 2 columns, and as many rows as scraped articles
#for each article save the keyword it is associated with and the articles content 
file_csv = open("BBC_scraped_data.csv", "w", encoding="utf-8", newline='')
csv_writer = csv.writer(file_csv)
csv_writer.writerow(["Keyword", "Content"])

#Use requests and beautifulsoup to scrape up to 100 links (or as many as possible) articles most relevant to keyword(s)
#then scrape these up to 100 links to collect article contents and save to a csv file
keyword_index = 0
for keyword_url in keyword_urls:
    pages_without = 0
    this_keyword_links = []
    article_count = 0
    page = 1
    retries = 0
    #only look through search result urls if less than 100 articles scraped
    #if unsuccessful request after 5 attempts move on to next keyword
    while article_count < 100 and retries <=5:
        try:
            #request resources from keyword_urls
            response = requests.get((keyword_url + str(page)), timeout=2.0)
            #raise a HTTPError if HTTP request has unsuccessful status code
            response.raise_for_status()
        #deal with HTTP errors
        except HTTPError as httperr:
            print(f"HTTP error: {httperr}")
            print("Retrying up to 5 times, currently on retry: ", retries+1)
            retries += 1
            continue
        #dealing with non HTTP errors
        except Exception as error:
            print(f"Other error: {error}")
            print("Retrying up to 5 times, currently on retry: ", retries+1)
            retries += 1
            continue

        #parse the unicode content of webpage with lxml parser
        parsed = BeautifulSoup((response.text), "lxml")

        #find each element in the list of results of the bbc keyword search
        #if no list of results, there are no more relevant results for this keyword, so exit the loop
        if parsed.find_all("div", class_="ssrcss-ut3rmk-Promo ett16tt0"):
            potential_article_links = parsed.find_all("div", class_="ssrcss-ut3rmk-Promo ett16tt0")
        else: 
            break
    
        #initialize list of links to relevant articles
        potential_articles = []
        for art in potential_article_links:
            #retrieve the link from results
            article_link = art.a.get("href")
            #discard artice if it is not a news article or not actually on the bbc website
            if("/news" not in article_link or "/sport" in article_link or "/newsround/" in article_link or 
            "/news-review" in article_link or "/live/" in article_link or "/blogs/" in article_link
            or "10159315" in article_link):
                continue
            else:
                potential_articles.append(article_link)
        
        #ensure only up to 100 links per keyword(s)
        if article_count >= 100:
            break
            
        #scrape articles in list
        cleaned_articles = []
        #use multithreading to speed up scraping of articles
        processes = []
        with ThreadPoolExecutor(max_workers=10) as executor:
            for potential_article in potential_articles:
                processes.append(executor.submit(scraper, potential_article))
        
        for task in as_completed(processes):
            cleaned_articles.append(task.result())

        article_count_before = article_count

        for article_s in cleaned_articles:
            if article_count > 99:
                break
            relev_test = article_s.lower()
            keyw_lower = keywords[keyword_index].lower()
            #ensure article is relevant by making sure it has the keyword or a variation of the keyword in
            #if we have not found an article with a variation of the keyword in 5 pages, stop being selective
            if (keyw_lower in relev_test or keyw_lower + "s" in relev_test or keyw_lower + "es" in relev_test
                or keyw_lower[:-3] + "ed" in relev_test or keyw_lower.split(" ")[0] in relev_test 
                or pages_without > 4):
                #write text of article to csv file
                csv_writer.writerow([keywords[keyword_index], article_s])
            else: 
                continue
            article_count += 1

        if article_count_before == article_count:
            pages_without += 1

        page += 1
        retries = 0
    
    keyword_index += 1

#close csv file
file_csv.close()

#print out how long it took the program to run
print("my program took ", time.time() - start_time, " seconds to run")


##problem 3 and 4 - takes about 40 secs to calculate similarity and 40 secs to visualize model (1 min 20 secs overall)
##I have combined these problems so it is easier to visualize the model I make with tSNE (explained in report)
##this first part of the code makes the model for semantic distances and creates a distance.xlsx file
from nltk.tokenize import word_tokenize 
from nltk.corpus import stopwords
import warnings 
#ignore warning about levenshtein distance submodule being disabled as we are not using this distance measure
warnings.filterwarnings(action='ignore', message = "The gensim.similarities.levenshtein submodule is disabled")  
from gensim.models import Word2Vec
import csv
import re
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from sklearn.manifold import TSNE
import requests
from requests import HTTPError
from bs4 import BeautifulSoup
import time

start_time = time.time()

def pre_process(iterable):
    #loop through articles and pre-process them
    word2vec_art = []
    for row in iterable:
        #convert to lowercase
        article_processed = row[1].lower()
        #remove special characters
        article_processed = re.sub("[^a-zA-Z]", " ", article_processed)
        #removed excess whitespace
        article_processed = re.sub(r"\s+", " ", article_processed)
        #word2vec works on single words, we will convert our multiword keywords into single words
        #e.g. dos attack -> dos_attack. Then anaylse the words around them to try and decipher semantic similarity
        #this will also convert any plural forms of keywords to the singular words e.g. targeted_threats -> targeted_threat
        article_processed = regular_exp.sub(lambda mo: keyword_dict[mo.string[mo.start():mo.end()]], article_processed)

        #create list of words for each article
        words = []
        for word in word_tokenize(article_processed):
            #remove stopwords
            if word in stop_words:
                continue
            else:
                words.append(word)
        
        word2vec_art.append(words)
    
    return word2vec_art

#list of stopwords
stop_words = stopwords.words()

#open excel file/workbook and create workbook object
filename = r"keywords.xlsx"
excel_file = openpyxl.load_workbook(filename)
excel_keywords = excel_file.active

#list of keywords
keywords = []

#loop through 1st column (except the title of the column) to read in keywords
for row_ in range(2, excel_keywords.max_row + 1):
    word = excel_keywords.cell(row = row_, column = 1)
    keywords.append(word.value)

#close excel file
excel_file.close()

#create dictionary for making keyword phrases into single words
keyword_dict = {}

for keyword in keywords:
    keyword = keyword.lower()
    spl = keyword.split(" ")
    out_word = ""
    if len(spl) > 1:
        for word in spl:
            out_word += word 
            if word != spl[-1]:
                out_word += "_"
    else:
        out_word = keyword

    if keyword[-1] == "s":
        plural_keyword = keyword + "es"
        keyword_dict[plural_keyword] = out_word
    elif keyword[-3:] != "ing":
        plural_keyword = keyword + "s"
        keyword_dict[plural_keyword] = out_word
    
    keyword_dict[keyword] = out_word

#regular expression for making keyword phrases into single words
regular_exp = re.compile("(%s)" % "|".join(map(re.escape, keyword_dict.keys())))

#create list of single word keywords
single_keywords = []
for keyw in keywords:
    s_key = regular_exp.sub(lambda mo: keyword_dict[mo.string[mo.start():mo.end()]], keyw.lower())
    single_keywords.append(s_key)

#list holding wikipedia article data
wiki_data = []

#get the wikipedia article for each keyword to increase data to train on (there isn't one for malicious bot)
base_w_url = "https://en.wikipedia.org/wiki/"
for keyword in keywords:
    keyword = regular_exp.sub(lambda mo: keyword_dict[mo.string[mo.start():mo.end()]], keyword.lower())
    keyword = keyword.capitalize()
    if keyword == "Malicious_bot":
        continue

    #request wikipedia data, if an error occurs retry up to 5 times before moving on to the next article
    article_retries = 0
    for _ in range(5):   
        try:
            #request resources from article url
            response = requests.get(base_w_url + keyword, timeout=2.0)
            #raise a HTTPError if HTTP request has unsucessful status code
            response.raise_for_status()
        #deal with HTTP errors
        except HTTPError as httperr:
            print(f"HTTP error: {httperr}")
            print("Retrying up to 5 times, currently on retry: ", article_retries+1)
            article_retries += 1
            continue
        #dealing with non HTTP errors
        except Exception as error:
            print(f"Other error: {error}")
            print("Retrying up to 5 times, currently on retry: ", article_retries+1)
            article_retries += 1
            continue
        break

    if article_retries > 4:
        continue
    
    wiki_str = ""

    #parse the unicode content of article with lxml parser
    parsed_wiki = BeautifulSoup((response.text), "lxml")

    #get heading
    heading = parsed_wiki.h1.text
    wiki_str += heading

    #get body, preserve order of text
    body = parsed_wiki.find("div", id="mw-content-text")
    texts = body.findAll(["p"])
    for text in texts:
        wiki_str += text.text + " \n"
    
    #save in wiki_data
    wiki_data.append(["wiki", wiki_str])

#open csv file in read mode
file_csv = open("BBC_scraped_data.csv", "r", newline="", encoding="utf-8")
csv_reader = csv.reader(file_csv)
next(csv_reader)

#pre-process articles to increase effectiveness of word2vec model
#initialise list of words in articles
word2vec_articles = []
wiki_artic = []

#pre-process bbc articles
word2vec_articles = pre_process(csv_reader)
#pre-process wiki articles
wiki_artic = pre_process(wiki_data)

word2vec_articles = word2vec_articles + wiki_artic

#create word2vec model
#we will use skip-gram as some of the words we are testing will not occur very frequently in our data
word2vec_model = Word2Vec(word2vec_articles, min_count = 1, window = 5, sg = 1, vector_size=50, epochs=10) 
#note "vector_size" argument is just "size" and "epochs" is "iter" in older Word2Vec installations

#use keywords.xlsx as a template to create distance.xlsx
file_excel = openpyxl.load_workbook("keywords.xlsx")
excel_workbook = file_excel.active
excel_workbook.title = "changed sheet"

#matrix storing distances to help with seaborn visualization in problem 4
distance_matrix = np.zeros((len(keywords), len(keywords)))

#use default cosine similarity to measure semantic distance, as this works well even for inputs of different size
#write this into our excel file
for row_ in range(2, excel_workbook.max_row + 1):
    for column_ in range(2, excel_workbook.max_column + 1):
        #read in keyword(s) and convert to single word form used in model
        word_1 = excel_workbook.cell(row = row_, column = 1).value.lower()
        word1 = regular_exp.sub(lambda mo: keyword_dict[mo.string[mo.start():mo.end()]], word_1)
        word_2 = excel_workbook.cell(row = 1, column = column_).value.lower()
        word2 = regular_exp.sub(lambda mo: keyword_dict[mo.string[mo.start():mo.end()]], word_2)
        cell = excel_workbook.cell(row = row_, column = column_)
        #write cosine similarity to excel file 
        similarity = word2vec_model.wv.similarity(word1, word2)
        cell.value = similarity
        distance_matrix[row_ - 2][column_ - 2] = similarity

#save new excel file
file_excel.save("distance.xlsx")

#close files
file_excel.close()
file_csv.close()

##this part of the code uses the model made to visualize keyword distance

#plot seaborn heatmap of semantic similarity between keywords
ax = sns.heatmap(distance_matrix, xticklabels = keywords, yticklabels = keywords, annot=True, cbar=False)
ax.figure.tight_layout()
plt.show()

#dimension reduction to 2 dimensions with TSNE to allow visualization of distances between keywords
#metric = cosine so cosine similarity is preserved
tsne_model = TSNE(perplexity=30, n_components=2, init="pca", n_iter=250, n_iter_without_progress=50, 
random_state=0, metric="cosine", square_distances=True)

#create list of words in the Word2Vec model and corresponding 50 dimensional vectors for them
#take note of the indexes of keywords
model_words = []
word_labels = []
keyword_labels = []
i = 0
for word in word2vec_model.wv.index_to_key:  #word2vec_model.wv.vocab in earlier gensim versions
    if word in single_keywords:
        keyword_labels.append(i)
    word_labels.append(word)
    model_words.append(word2vec_model.wv[word])
    i += 1

#reduce 50 dimensional vectors to 2 dimensional vectors
values_2d = tsne_model.fit_transform(model_words)

#visualize the distance between 2 keywords with a scatter plot
key1 = keyword_labels[0]
key2 = keyword_labels[1]

#append x and y coordinates of 2 dimensional keyword vectors to x_ and y_
x_ = []
y_ = []
x_.append(values_2d[key1][0])
x_.append(values_2d[key2][0])
y_.append(values_2d[key1][1])
y_.append(values_2d[key2][1])

#create scatter plot with seaborn
scat = sns.scatterplot(x=x_, y=y_)

#annotate points on plot
plt.annotate(word_labels[key1],xy=(x_[0], y_[0]), xytext=(-30, 8), textcoords='offset points')
plt.annotate(word_labels[key2],xy=(x_[1], y_[1]), xytext=(7, -4), textcoords='offset points')

#show graph
plt.show()

#visualize distance between all keywords at once
x__ = []
y__ = []

#append x and y coordinates of 2 dimensional keyword vectors to x__ and y__
for index in keyword_labels:
    x__.append(values_2d[index][0])
    y__.append(values_2d[index][1])

#create scatter plot with seaborn
scatter = sns.scatterplot(x=x__, y=y__)

#label points
j = 0
for label_ind in keyword_labels:
    if j == 0:
        print(x__[j], y__[j])
    if word_labels[label_ind] == "targeted_threat":
        plt.annotate(word_labels[label_ind], xy=(x__[j], y__[j]), xytext=(7, -4), textcoords='offset points')
    else:
        plt.annotate(word_labels[label_ind], xy=(x__[j], y__[j]), xytext=(-20, 5), textcoords='offset points') 
    j += 1

#show graph
plt.show()

#print out how long it took the program to run
print("my program took ", time.time() - start_time, " seconds to run") 
